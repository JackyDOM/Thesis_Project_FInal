from flask import Flask, render_template, request, send_file, send_from_directory, jsonify
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
import openpyxl
from flask_cors import CORS
import os
import random
import json
import uuid
import string
import re
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
import re

app = Flask(__name__)
CORS(app)

# Register the basename filter
from os.path import basename
app.jinja_env.filters['basename'] = basename


# Ensure an 'uploads' directory exists
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Ensure 'static/images' directory exists
STATIC_IMAGES_FOLDER = os.path.join('static', 'images')
if not os.path.exists(STATIC_IMAGES_FOLDER):
    os.makedirs(STATIC_IMAGES_FOLDER)

UPLOAD_FOLDER = os.path.join('static', 'uploads')  # Path to the uploads folder
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create uploads folder if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Initialize the database (create table if it doesn’t exist)
def init_db():
    with sqlite3.connect('database.db') as con:
        cur = con.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS students (
                first_name TEXT,
                second_name TEXT,
                age INTEGER,
                gender TEXT,
                email TEXT,
                school_name TEXT,
                addr TEXT,
                city TEXT,
                filename TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                policy_id TEXT,
                sel TEXT,
                tran_date TEXT,
                time TEXT,
                code TEXT,
                description TEXT,
                loc TEXT
            )
        """)
        con.commit()


import sqlite3

con = sqlite3.connect("database.db")
con.row_factory = sqlite3.Row
cur = con.cursor()
cur.execute("SELECT rowid, * FROM students")
rows = cur.fetchall()
print("All records:")
for row in rows:
    print(dict(row))
con.close()

# Call this once to set up the database
init_db()

# Home Page route
@app.route("/")
def home():
    return render_template("home.html")

# Route to form used to add a new student to the database
@app.route("/enternew")
def enternew():
    return render_template("student.html")

def is_english_only(text):
    try:
        text.encode('ascii')
        return True
    except UnicodeEncodeError:
        return False

@app.route("/addrec", methods=['POST', 'GET'])
def addrec():
    if request.method == 'POST':
        try:
            # Get data from the form
            first_name = request.form['first_name']
            second_name = request.form['second_name']
            age = request.form['age']
            gender = request.form['gender']
            email = request.form['email']
            school_name = request.form['school_name']
            addr = request.form['add']
            city = request.form['city']

            # Validation code (unchanged)
            def has_special_characters(text):
                pattern = r'^[a-zA-Z0-9\s.,-]+$'
                return not bool(re.match(pattern, text))

            for field_name, value in {
            'First Name': first_name,
            'Second Name': second_name,
            'School Name': school_name,
            'Address': addr,
            'City': city
            }.items():
                if has_special_characters(value):
                    return render_template('result.html', data={"error": f"English only is allowed in this field."})
                if not is_english_only(value):
                    return render_template('result.html', data={"error": f"English only is allowed in this field."})

            if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                return render_template('result.html', data={"error": "Invalid email format"})

            # Handle image upload
            image_filename = None
            if 'profile_image' in request.files:
                file = request.files['profile_image']
                if file and file.filename != '':
                    # Validate file extension
                    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
                    if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                        # Use the original filename
                        image_filename = file.filename
                        file.save(os.path.join(app.config['UPLOAD_FOLDER'], image_filename))
                    else:
                        return render_template('result.html', data={"error": "Invalid image format. Allowed: png, jpg, jpeg, gif"})

            # Insert into database
            with sqlite3.connect('database.db') as con:
                cur = con.cursor()
                cur.execute("""
                    INSERT INTO students 
                    (first_name, second_name, age, gender, email, school_name, addr, city, filename) 
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (first_name, second_name, age, gender, email, school_name, addr, city, image_filename))
                policy_id = cur.lastrowid
                policy_id_str = f"POL{policy_id:05d}"

                submission_date = datetime.now().strftime("%d/%m/%Y")
                submission_time = datetime.now().strftime("%H:%M:%S")
                
                def generate_random_sel():
                    return f"{random.randint(10000, 99999):05d}"

                def generate_random_code():
                    return f"B{random.randint(100, 999)}"
                
                transactions = [
                    {
                        "sel": generate_random_sel(),
                        "tran_date": submission_date,
                        "time": submission_time,
                        "code": generate_random_code(),
                        "description": f"Student Registration - {first_name} {second_name}",
                        "loc": "P"
                    },
                    {
                        "sel": generate_random_sel(),
                        "tran_date": submission_date,
                        "time": submission_time,
                        "code": generate_random_code(),
                        "description": f"School Name - {school_name}",
                        "loc": "P"
                    }
                ]
                for tran in transactions:
                    cur.execute("""
                        INSERT INTO transactions (policy_id, sel, tran_date, time, code, description, loc)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (policy_id_str, tran["sel"], tran["tran_date"], tran["time"], tran["code"], tran["description"], tran["loc"]))
                con.commit()

            # Update students.xlsx
            file_path = os.path.join(app.static_folder, 'students.xlsx')
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.append(['First Name', 'Second Name', 'Age', 'Gender', 'Email', 'School Name', 'Address', 'City', 
                           'Image Filename', 'SEL', 'Tran Date', 'Time', 'Code', 'Description', 'Loc'])
            else:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

            # Write one row, including image filename
            tran = transactions[0]
            ws.append([first_name, second_name, age, gender, email, school_name, addr, city,
                       image_filename if image_filename else '',
                       tran["sel"], tran["tran_date"], tran["time"], tran["code"], school_name, tran["loc"]])
            wb.save(file_path)

            detail_data = [
                {
                    "first_name": first_name,
                    "second_name": second_name,
                    "age": age,
                    "gender": gender,
                    "email": email,
                    "school_name": school_name,
                    "addr": addr,
                    "city": city,
                    "image_filename": image_filename if image_filename else '',
                    "sel": transactions[0]["sel"],
                    "tran_date": transactions[0]["tran_date"],
                    "time": transactions[0]["time"],
                    "code": transactions[0]["code"],
                    "description": transactions[0]["description"],
                    "loc": transactions[0]["loc"]
                }
            ]

            data = {
                "policy_id": policy_id_str,
                "first_name": first_name,
                "second_name": second_name,
                "city": city,
                "image_filename": image_filename if image_filename else '',
                "transactions": transactions,
                "detail_data": detail_data
            }
            return render_template('result.html', data=data)

        except Exception as e:
            if 'con' in locals():
                con.rollback()
            msg = f"Error in the INSERT: {str(e)}"
            return render_template('result.html', data={"error": msg})
        finally:
            if 'con' in locals():
                con.close()

@app.route("/run", methods=['POST'])
def run():
    try:
        rowid = request.form['id']
        with sqlite3.connect("database.db") as con:
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            cur.execute("SELECT * FROM students WHERE rowid = ?", (rowid,))
            student = cur.fetchone()
            if not student:
                return render_template('result.html', data={"error": "Student not found"})

            policy_id = f"POL{int(rowid):05d}"
            cur.execute("SELECT sel, tran_date, time, code, description, loc FROM transactions WHERE policy_id = ?", (policy_id,))
            transactions = [{"sel": row[0], "tran_date": row[1], "time": row[2], "code": row[3], "description": row[4], "loc": row[5]} for row in cur.fetchall()]
            
            image_filename = student['filename'] if student['filename'] else ''

            detail_data = [
                {
                    "first_name": student["first_name"],
                    "second_name": student["second_name"],
                    "age": student["age"],
                    "gender": student["gender"],
                    "email": student["email"],
                    "school_name": student["school_name"],
                    "addr": student["addr"],
                    "city": student["city"],
                    "image_filename": image_filename,
                    "sel": transactions[0]["sel"],
                    "tran_date": transactions[0]["tran_date"],
                    "time": transactions[0]["time"],
                    "code": transactions[0]["code"],
                    "description": transactions[0]["description"],
                    "loc": transactions[0]["loc"]
                }
            ] if transactions else []

        data = {
            "policy_id": policy_id,
            "first_name": student["first_name"],
            "second_name": student["second_name"],
            "city": student["city"],
            "image_filename": image_filename,
            "transactions": transactions,
            "detail_data": detail_data
        }
        return render_template('result.html', data=data)

    except Exception as e:
        return render_template('result.html', data={"error": f"Error in RUN: {str(e)}"})
    

# Route to list students (GET request for viewing students)
@app.route('/student_list', methods=['GET'])
def student_list():
    con = sqlite3.connect("database.db")
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute("SELECT rowid, * FROM students")
    rows = cur.fetchall()
    con.close()
    return render_template("list.html", rows=rows)

# Load DeepSeek model and tokenizer
model_path = "./deepseek_r1_model"
tokenizer = AutoTokenizer.from_pretrained(model_path)
model = AutoModelForCausalLM.from_pretrained(model_path, torch_dtype=torch.float32).to("cpu")
model.eval()

def query_deepseek(prompt):
    inputs = tokenizer(prompt, return_tensors="pt").to("cpu")
    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=50,
            pad_token_id=tokenizer.eos_token_id,
            eos_token_id=tokenizer.eos_token_id,
            do_sample=False
        )
   
    generated_ids = outputs[0][inputs['input_ids'].shape[-1]:]
    result = tokenizer.decode(generated_ids, skip_special_tokens=True).strip()
    print(f"[Raw DeepSeek Output] {result}")
    return result

def normalize_email(email):
    """Remove spaces from email addresses."""
    return re.sub(r'\s+', '', email)

# Route to handle the form submission (POST request)
@app.route('/submit_student', methods=['POST'])
def submit_student():
    # Step 1: Get data from frontend (multipart/form-data)
    try:
        first_name = request.form.get('first_name', '').strip().lower()
        second_name = request.form.get('second_name', '').strip().lower()
        age = request.form.get('age', '').strip()
        gender = request.form.get('gender', '').strip().lower()
        email = normalize_email(request.form.get('email', '').strip().lower())
        school_name = request.form.get('school_name', '').strip().lower()
        address = request.form.get('address', request.form.get('addr', '')).strip().lower()
        city = request.form.get('city', '').strip().lower()

        image = request.files.get('image')
        image_name = image.filename if image else None

        if not all([first_name, second_name, age, gender, email]):
            print(f"[Validation Error] Missing required fields")
            return jsonify({"error": "All required fields must be filled."}), 400

        try:
            age = int(age)
        except ValueError:
            print(f"[Validation Error] Invalid age: {age}")
            return jsonify({"error": "Age must be a valid number."}), 400

        print(f"[Frontend Data] {first_name}, {second_name}, {age}, {gender}, {email}, {school_name}, {address}, {city}, {image_name}")

    except Exception as e:
        print(f"[Error] Failed to parse form data: {str(e)}")
        return jsonify({"error": "Invalid form data."}), 400

    # Step 2: Read and process OCR JSON data
    ocr_records = []
    ocr_json_path = os.path.join('ocr', 'extract_ocr', 'ocr_output.json')
    try:
        if os.path.exists(ocr_json_path):
            with open(ocr_json_path, 'r', encoding='utf-8') as f:
                ocr_json_data = json.load(f)
            
            rows_by_y = {}
            for entry in ocr_json_data:
                y_top = entry.get('bounding_box', [[0, 0]])[0][1]
                y_key = round(y_top / 10)
                if y_key not in rows_by_y:
                    rows_by_y[y_key] = []
                rows_by_y[y_key].append(entry)
            
            sorted_y_keys = sorted(rows_by_y.keys())
            field_order = ['First Name', 'Second Name', 'Age', 'Gender', 'Email', 'School Name', 'Address', 'City', 'Result']
            expected_fields = ['First Name', 'Second Name', 'Age', 'Gender', 'Email']
            header_row = None

            for y_key in sorted_y_keys:
                row_entries = sorted(rows_by_y[y_key], key=lambda x: x.get('bounding_box', [[0, 0]])[0][0])
                row_texts = [entry.get('text', '').strip() for entry in row_entries]
                print(f"[Processing OCR Row] y={y_key}, texts={row_texts}")

                if not header_row and any(text.lower() in [field.lower() for field in field_order] for text in row_texts):
                    header_row = row_texts
                    print(f"[Header Row Detected] {header_row}")
                    continue

                if len(row_texts) < len(expected_fields):
                    print(f"[Skipping Row] Insufficient fields: {row_texts}")
                    continue

                current_record = {}
                for i, text in enumerate(row_texts):
                    if i < len(field_order):
                        current_record[field_order[i]] = text
                
                if all(field in current_record for field in expected_fields):
                    try:
                        age_text = current_record['Age'].strip()
                        int(age_text)
                        ocr_records.append(current_record)
                        print(f"[Added OCR Record] {current_record}")
                    except ValueError:
                        print(f"[Skipping Record] Invalid age value: {age_text}")
                        continue
                else:
                    print(f"[Skipping Row] Missing required fields: {current_record}")

            print(f"[OCR Records] {ocr_records}")
        else:
            print(f"[Warning] OCR JSON file not found at {ocr_json_path}")
            return jsonify({"error": "OCR data not found."}), 400
    except Exception as e:
        print(f"[Error] Failed to read OCR JSON: {str(e)}")
        return jsonify({"error": "Failed to process OCR data."}), 500

    # Step 3: Compare frontend input with OCR records using DeepSeek
    field_comparisons = []
    match_found = False
    matched_record = None
    row_index = 0

    for ocr_record in ocr_records:
        row_index += 1
        try:
            ocr_first_name = ocr_record.get('First Name', '').strip().lower()
            ocr_second_name = ocr_record.get('Second Name', '').strip().lower()
            ocr_age = int(ocr_record.get('Age', '0').strip())
            ocr_gender = ocr_record.get('Gender', '').strip().lower()
            ocr_email = normalize_email(ocr_record.get('Email', '').strip().lower())

            ocr_text_segment = f"{ocr_first_name} {ocr_second_name} {ocr_age} {ocr_gender} {ocr_email}".strip()
            input_text_segment = f"{first_name} {second_name} {age} {gender} {email}".strip()

            prompt = (
                f"Compare the following input and OCR record field by field:\n\n"
                f"Input: {input_text_segment}\n"
                f"OCR Record: {ocr_text_segment}\n\n"
                f"Return ONLY 'MATCH' or 'NOT MATCH' in uppercase. "
                f"Return 'MATCH' only if ALL fields (First Name, Second Name, Age, Gender, Email) match exactly "
                f"with case-insensitive string comparison, exact number match, and emails normalized by removing spaces. "
                f"Return 'NOT MATCH' otherwise. "
                f"Do NOT add explanations or extra text. Output only MATCH or NOT MATCH."
            )

            print(f"[Checking Row {row_index}] {ocr_text_segment}")
            print(f"[DeepSeek Prompt for OCR] {prompt}")
            response = query_deepseek(prompt)
            print(f"[Raw DeepSeek Output] {response}")

            # Normalize DeepSeek output to 'MATCH' or 'NOT MATCH'
            response_upper = response.strip().upper()
            match = re.search(r'\b(MATCH|NOT MATCH)\b', response_upper)
            normalized_response = match.group(0) if match else "NOT MATCH"
            print(f"[Normalized DeepSeek Response] {normalized_response}")

            ocr_match = normalized_response == "MATCH"

            if ocr_match:
                comparisons = [
                    {"field": "first_name", "backend_value": ocr_first_name, "frontend_value": first_name, "result": "Pass"},
                    {"field": "second_name", "backend_value": ocr_second_name, "frontend_value": second_name, "result": "Pass"},
                    {"field": "age", "backend_value": ocr_age, "frontend_value": age, "result": "Pass"},
                    {"field": "gender", "backend_value": ocr_gender, "frontend_value": gender, "result": "Pass"},
                    {"field": "email", "backend_value": ocr_email, "frontend_value": email, "result": "Pass"},
                ]

                field_comparisons = [{
                    "row_id": row_index,
                    "ocr_record": ocr_record,
                    "comparisons": comparisons,
                    "deepseek": {
                        "prompt": prompt,
                        "response": response,
                        "normalized_response": normalized_response,
                        "ocr_match": ocr_match
                    }
                }]

                match_found = True
                matched_record = {
                    "first_name": ocr_first_name,
                    "second_name": ocr_second_name,
                    "age": ocr_age,
                    "gender": ocr_gender,
                    "email": ocr_email,
                    "school_name": ocr_record.get('School Name', '').lower(),
                    "address": ocr_record.get('Address', '').lower(),
                    "city": ocr_record.get('City', '').lower(),
                    "result": ocr_record.get('Result', '').lower(),
                    "image_name": image_name
                }
                print(f"[Match Found] Row {row_index} matched: {matched_record}")
                break

        except Exception as e:
            print(f"[OCR Processing Error] Skipping row {row_index}: {str(e)}")
            continue



    # Step 4: Return response
    print(f"[Final Response] match_found: {match_found}")
    if match_found:
        return jsonify({
            "message": "Match found!",
            "match": True,
            "student": matched_record,
            "field_comparisons": field_comparisons
        }), 200
    else:
        return jsonify({
            "message": "No match found.",
            "match": False,
            "student": None,
            "field_comparisons": []
        }), 200

    # Step 5: No match found
    # print("[No Match] Returning no match")
    # mismatch_fields = []
    # if rows:
    #     db_data = {
    #         "first_name": rows[-1]["first_name"].strip().lower(),
    #         "second_name": rows[-1]["second_name"].strip().lower(),
    #         "age": int(rows[-1]["age"]),
    #         "gender": rows[-1]["gender"].strip().lower(),
    #         "email": rows[-1]["email"].strip().lower()
    #     }
    #     mismatch_fields = [
    #         field for field, input_val, db_val in [
    #             ("first_name", first_name.lower(), db_data["first_name"]),
    #             ("second_name", second_name.lower(), db_data["second_name"]),
    #             ("age", age, db_data["age"]),
    #             ("gender", gender.lower(), db_data["gender"]),
    #             ("email", email.lower(), db_data["email"])
    #         ] if input_val != db_val
    #     ]

    # Include OCR match result in the no-match response
    # ocr_match = deepseek_data.get("ocr_match", False) if deepseek_data else False
    # return jsonify({
    #     "message": "No match found.",
    #     "match": False,
    #     "student": {
    #         "first_name": first_name,
    #         "second_name": second_name,
    #         "age": age,
    #         "gender": gender,
    #         "email": email,
    #         "school_name": school_name if school_name else 'N/A',
    #         "address": address if address else 'N/A',
    #         "city": city if city else 'N/A',
    #         "image_name": image_name if image_name else 'No Image',
    #         "ocr_match": ocr_match
    #     },
    #     "mismatch_fields": mismatch_fields,
    #     "field_comparisons": field_comparisons
    # }), 200

    

@app.route("/edit", methods=['POST', 'GET'])
def edit():
    if request.method == 'POST':
        try:
            id = request.form['id']
            con = sqlite3.connect("database.db")
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            cur.execute("SELECT rowid, * FROM students WHERE rowid = ?", (id,))
            rows = cur.fetchall()
        except:
            rows = []
        finally:
            con.close()
            return render_template("edit.html", rows=rows)

@app.route("/editrec", methods=['POST', 'GET'])
def editrec():
    if request.method == 'POST':
        try:
            rowid = request.form['rowid']
            first_name = request.form['first_name']
            second_name = request.form['second_name']
            age = request.form['age']
            gender = request.form['gender']
            email = request.form['email']
            school_name = request.form['school_name']
            addr = request.form['add']
            city = request.form['city']

            # Update the database
            with sqlite3.connect('database.db') as con:
                cur = con.cursor()
                cur.execute("""
                    UPDATE students 
                    SET first_name = ?, second_name = ?, age = ?, gender = ?, email = ?, 
                        school_name = ?, addr = ?, city = ?
                    WHERE rowid = ?
                """, (first_name, second_name, age, gender, email, school_name, addr, city, rowid))
                policy_id = f"POL{int(rowid):05d}"
                # Fetch existing transactions for this student
                cur.execute("SELECT sel, tran_date, time, code, description, loc FROM transactions WHERE policy_id = ?", (policy_id,))
                transactions = [{"sel": row[0], "tran_date": row[1], "time": row[2], "code": row[3], "description": row[4], "loc": row[5]} for row in cur.fetchall()]
                con.commit()

            # Update students.xlsx with only one row per student
            file_path = os.path.join(app.static_folder, 'students.xlsx')
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Remove old rows for this student (match by name as a simple approach)
            rows_to_keep = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] != first_name or row[1] != second_name:
                    rows_to_keep.append(row)

            # Rewrite the Excel file
            ws.delete_rows(1, ws.max_row)  # Clear all rows
            ws.append(['First Name', 'Second Name', 'Age', 'Gender', 'Email', 'School Name', 'Address', 'City', 
                       'SEL', 'Tran Date', 'Time', 'Code', 'Description', 'Loc'])
            for row in rows_to_keep:
                ws.append(row)

            # Add updated student data with one row, using the first transaction
            tran = transactions[0]  # Use the first transaction
            ws.append([first_name, second_name, age, gender, email, school_name, addr, city,
                       tran["sel"], tran["tran_date"], tran["time"], tran["code"], school_name, tran["loc"]])
            wb.save(file_path)

            msg = "Record successfully edited in the database and Excel file"
            data = {"message": msg}

        except Exception as e:
            if 'con' in locals():
                con.rollback()
            msg = f"Error in the Edit: {str(e)}"
            data = {"error": msg}
        finally:
            if 'con' in locals():
                con.close()
            return render_template('result.html', data=data)

@app.route("/delete", methods=['POST', 'GET'])
def delete():
    if request.method == 'POST':
        try:
            rowid = request.form['id']
            policy_id = f"POL{int(rowid):05d}"
            file_path = os.path.join(app.static_folder, 'students.xlsx')

            # Delete from database and handle image file
            with sqlite3.connect('database.db') as con:
                con.row_factory = sqlite3.Row
                cur = con.cursor()
                # Fetch the filename before deleting
                cur.execute("SELECT filename FROM students WHERE rowid = ?", (rowid,))
                student = cur.fetchone()
                if not student:
                    return render_template('result.html', data={"error": "Student not found"})

                # Delete the image file if it exists
                image_filename = student['filename']
                if image_filename:
                    image_path = os.path.join(app.static_folder, 'uploads', image_filename)
                    if os.path.exists(image_path):
                        os.remove(image_path)

                # Delete student and transaction records
                cur.execute("DELETE FROM students WHERE rowid = ?", (rowid,))
                cur.execute("DELETE FROM transactions WHERE policy_id = ?", (policy_id,))
                con.commit()

                # Fetch remaining students and their transactions
                cur.execute("SELECT first_name, second_name, age, gender, email, school_name, addr, city, filename, rowid FROM students")
                students = cur.fetchall()

            # Rewrite students.xlsx with one row per student
            wb = Workbook()
            ws = wb.active
            ws.append(['First Name', 'Second Name', 'Age', 'Gender', 'Email', 'School Name', 'Address', 'City', 
                       'Image Filename', 'SEL', 'Tran Date', 'Time', 'Code', 'Description', 'Loc'])

            for student in students:
                first_name, second_name, age, gender, email, school_name, addr, city, filename, student_rowid = student
                student_policy_id = f"POL{student_rowid:05d}"
                cur.execute("SELECT sel, tran_date, time, code, description, loc FROM transactions WHERE policy_id = ?", (student_policy_id,))
                transactions = cur.fetchall()
                if transactions:  # Only append if there are transactions
                    tran = transactions[0]  # Use the first transaction
                    ws.append([first_name, second_name, age, gender, email, school_name, addr, city,
                              filename if filename else '', tran[0], tran[1], tran[2], tran[3], school_name, tran[5]])

            wb.save(file_path)

            data = {"message": "Record and associated image successfully deleted from the database and Excel file"}
        except Exception as e:
            if 'con' in locals():
                con.rollback()
            data = {"error": f"Error in the DELETE: {str(e)}"}
        finally:
            if 'con' in locals():
                con.close()
            return render_template('result.html', data=data)


# when user upload file and click on submit so it go to displayverify.html file
from PIL import Image, ImageDraw, ImageFont
import os
import uuid
import textwrap

def generate_five_images():
    font_size = 20
    row_padding = 40
    header_padding = 60
    col_spacing = 60
    margin_left = 20
    margin_top = 20

    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except IOError:
        font = ImageFont.load_default()

    # Connect to database
    with sqlite3.connect("database.db") as con:
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        cur.execute("""SELECT first_name, second_name, age, gender, email, 
                       school_name, addr, city FROM students""")
        db_rows = cur.fetchall()

    if not db_rows:
        raise ValueError("No data found in the database.")

    all_results = [dict(row) for row in db_rows]
    original_headers = list(all_results[0].keys())

    dummy_img = Image.new("RGB", (1, 1))
    draw = ImageDraw.Draw(dummy_img)

    # Output folder
    output_folder = os.path.join("static", "generate_test_case")
    os.makedirs(output_folder, exist_ok=True)

    row_height = int(font_size * 1.5)
    content_height = header_padding + (row_height + row_padding) * len(all_results)
    total_height = margin_top + content_height + 20

    # Generate images with rotated column headers
    for i in range(10):
        # Rotate headers
        headers = original_headers[i:] + original_headers[:i]

        # Calculate column widths based on rotated headers
        col_widths = []
        for header in headers:
            max_width = draw.textbbox((0, 0), header, font=font)[2]
            for row in all_results:
                text_width = draw.textbbox((0, 0), str(row[header]), font=font)[2]
                max_width = max(max_width, text_width)
            col_widths.append(max_width + 10)

        total_width = sum(col_widths) + (len(headers) - 1) * col_spacing + 2 * margin_left

        # Create image
        img = Image.new("RGB", (total_width, total_height), color="white")
        draw = ImageDraw.Draw(img)

        # Draw headers
        x_offset = margin_left
        y_offset = margin_top
        for j, header in enumerate(headers):
            draw.text((x_offset, y_offset), header, font=font, fill="black")
            x_offset += col_widths[j] + col_spacing
        y_offset += header_padding

        # Draw data rows
        for row in all_results:
            x_offset = margin_left
            for j, header in enumerate(headers):
                draw.text((x_offset, y_offset), str(row[header]), font=font, fill="black")
                x_offset += col_widths[j] + col_spacing
            y_offset += row_height + row_padding

        # Save image
        image_filename = f"students_rotated_{i + 1}.png"
        image_path = os.path.join(output_folder, image_filename)
        img.save(image_path)
        print(f"Image saved: {image_path}")


def create_image_from_results(results):
    # Adjusted parameters to match the image
    font_size = 20  # Reduced font size for a closer match
    row_padding = 50  # Tighter vertical spacing between rows
    header_padding = 60  # Reduced space between header and first row
    col_spacing = 60  # Slightly reduced horizontal spacing between columns
    margin_left = 20  # Reduced left margin
    margin_top = 20  # Reduced top margin

    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except IOError:
        font = ImageFont.load_default()

    dummy_img = Image.new("RGB", (1, 1))
    draw = ImageDraw.Draw(dummy_img)

    headers = list(results[0].keys())
    num_rows = len(results)

    # Calculate column widths
    col_widths = []
    for header in headers:
        max_width = draw.textbbox((0, 0), header, font=font)[2]
        for row in results:
            value = str(row[header])
            text_width = draw.textbbox((0, 0), value, font=font)[2]
            max_width = max(max_width, text_width)
        col_widths.append(max_width + 10)  # Reduced padding per column

    # Calculate total width and height
    total_width = sum(col_widths) + (len(headers) - 1) * col_spacing + 2 * margin_left
    row_height = int(font_size * 1.5)  # Adjusted row height for tighter spacing
    estimated_height = margin_top + header_padding + (row_height + row_padding) * num_rows + 20  # Reduced extra padding

    img = Image.new("RGB", (total_width, estimated_height), color="white")
    draw = ImageDraw.Draw(img)

    # Draw headers
    x_offset = margin_left
    y_offset = margin_top
    for i, header in enumerate(headers):
        draw.text((x_offset, y_offset), header, font=font, fill="black")
        x_offset += col_widths[i] + col_spacing
    y_offset += header_padding

    # Draw rows
    for row in results:
        x_offset = margin_left
        for i, header in enumerate(headers):
            text = str(row[header])
            draw.text((x_offset, y_offset), text, font=font, fill="black")
            x_offset += col_widths[i] + col_spacing
        y_offset += row_height + row_padding

    # Create folder if not exists
    output_folder = "static/images"
    os.makedirs(output_folder, exist_ok=True)

    image_path = os.path.join(output_folder, f"{uuid.uuid4().hex}.png")
    img.save(image_path)

    return image_path


@app.route("/upload", methods=["GET", "POST"], endpoint="upload")
def upload_file():
    if request.method == "POST":
        try:
            if "file" not in request.files:
                return render_template("upload.html", error="No file uploaded")
            
            file = request.files["file"]
            if file.filename == "":
                return render_template("upload.html", error="No file selected")
            if not file.filename.lower().endswith(".xlsx"):
                return render_template("upload.html", error="Only .xlsx files are allowed!")
            
            # Save the uploaded file
            unique_filename = f"{uuid.uuid4().hex}_{file.filename}"
            upload_path = os.path.join(UPLOAD_FOLDER, unique_filename)
            file.save(upload_path)
            
            # Process the file
            df_uploaded = pd.read_excel(upload_path)
            
            # Fetch data from the database
            with sqlite3.connect("database.db") as con:
                con.row_factory = sqlite3.Row
                cur = con.cursor()
                cur.execute("""SELECT first_name, second_name, age, gender, email, 
                               school_name, addr, city FROM students""")
                db_rows = cur.fetchall()
            
            db_df = pd.DataFrame(db_rows, columns=[
                'First Name', 'Second Name', 'Age', 'Gender', 'Email',
                'School Name', 'Address', 'City'
            ])
            
            # Prepare comparison results
            comparison_results = []
            student_columns = ['First Name', 'Second Name', 'Age', 'Gender', 'Email', 'School Name', 'Address', 'City']
            
            # Compare the data
            for index, uploaded_row in df_uploaded.iterrows():
                matching_rows = db_df[
                    (db_df['First Name'] == uploaded_row['First Name']) &
                    (db_df['Second Name'] == uploaded_row['Second Name'])
                ]
                
                row_data = {col: uploaded_row[col] for col in student_columns}
                
                if not matching_rows.empty:
                    match_row = matching_rows.iloc[0]
                    is_match = True
                    for col in student_columns:
                        if str(uploaded_row[col]) != str(match_row[col]):
                            is_match = False
                    row_data['Result'] = 'Pass' if is_match else 'Fail'
                else:
                    row_data['Result'] = 'Fail'
                
                comparison_results.append(row_data)
            
            # Generate image from results
            image_path = create_image_from_results(comparison_results)
            
            # Clean up: remove the temporary uploaded file
            os.remove(upload_path)
            
            return render_template("verifyResult.html", 
                                   results=comparison_results,
                                   message="Verification completed",
                                   image_path=image_path)
        
        except Exception as e:
            return render_template("upload.html", error=f"Error processing file: {str(e)}")
    
    return render_template("upload.html")


# New route to serve the generated Excel file
@app.route("/download/<filename>")
def download_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


# for route Result
@app.route("/displayResult")
def displayResult():
    return render_template("displayResult.html")

# if __name__ == "__main__":
#     # generate_five_images()
#     serve(app, host='127.0.0.1', port=5000)